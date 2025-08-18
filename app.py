from flask import Flask, render_template, request, send_from_directory, redirect, url_for, session, abort
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import uuid
import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
import matplotlib.patches as patches
from flask_mail import Mail, Message
import requests
from flask_session import Session
import re
from io import BytesIO
from flask import send_file, abort, current_app



# ================== UYGULAMA & OTURUM ==================
app = Flask(__name__)
app.secret_key = 'gizli_anahtar'
app.permanent_session_lifetime = timedelta(minutes=10)
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# ================== MAİL ==================
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'winJohariTest@gmail.com'
app.config['MAIL_PASSWORD'] = 'pebxhabcyhcucsmy'
app.config['MAIL_DEFAULT_SENDER'] = 'winJohariTest@gmail.com'
mail = Mail(app)

# ================== DİZİNLER & AYARLAR ==================
BASE_DIR = os.getcwd()
STATIC_DIR = os.path.join(BASE_DIR, "static")
GRAFIK_DIR = os.path.join(STATIC_DIR, "grafik")
DATA_DIR = os.path.join(BASE_DIR, "data")  # gizli tutulacak
os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(GRAFIK_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

EXCEL_YOLU = os.path.join(DATA_DIR, "sonuclar.xlsx")  # artık static'te değil!
ADMIN_PASS = os.environ.get("ADMIN_PASS", "admin123")  # basit indirme koruması

# Güvenli dosya adı
def slugify(txt):
    txt = re.sub(r"[^\w\s-]", "", str(txt), flags=re.UNICODE).strip()
    txt = re.sub(r"[\s-]+", "_", txt)
    return txt[:40] if txt else "kisi"


def _harfe_cevir(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    if s in ("A","B","C","D"):
        return s
    if s in ("1","2","3","4"):
        return {"1":"A","2":"B","3":"C","4":"D"}[s]
    return None  # tanınmayan değer
from openpyxl import load_workbook
from io import BytesIO
from flask import send_file, abort, current_app, flash

def _exceli_oku_ve_hesapla(file_storage):
    """
    Beklenen başlıklar: Yapan, Test Edilen, S1 .. S48
    Cevaplar A/B/C/D ya da 1/2/3/4 olabilir.
    Dönüş: (sonuclar_listesi, hatalar_listesi)
    """
    wb = load_workbook(filename=file_storage, data_only=True)
    ws = wb.active

    # Başlık satırı
    headers = [ (c.value or "").strip() if isinstance(c.value,str) else c.value for c in ws[1] ]
    # Esnek: Türkçe büyük/küçük harf farkını yumuşat
    h_lower = [ (h or "").lower() for h in headers ]

    def _idx(name):
        name_l = name.lower()
        if name_l in h_lower:
            return h_lower.index(name_l)
        return None

    yapan_idx = _idx("yapan")
    edilen_idx = _idx("test edilen")
    # S1..S48 sütunları
    s_idx = []
    for i in range(1,49):
        key = f"s{i}"
        j = _idx(key)
        if j is None:
            # Büyük harfe duyarsız ara: S1 / s1
            try:
                j = [x.lower() for x in headers].index(key)
            except ValueError:
                j = None
        s_idx.append(j)

    hatalar = []
    sonuclar = []
    if yapan_idx is None or edilen_idx is None or any(j is None for j in s_idx):
        hatalar.append("Başlıklar eksik. 'Yapan', 'Test Edilen', 'S1'..'S48' olmalı.")
        return [], hatalar

    # Satırları işle
    for row_idx in range(2, ws.max_row + 1):
        yapan = ws.cell(row=row_idx, column=yapan_idx+1).value
        edilen = ws.cell(row=row_idx, column=edilen_idx+1).value
        if not yapan and not edilen:
            continue  # boş satır

        # 48 cevap topla
        cevaplar = {}
        eksik = []
        for i in range(1,49):
            col = s_idx[i-1] + 1
            ham = ws.cell(row=row_idx, column=col).value
            harf = _harfe_cevir(ham)
            if not harf:
                eksik.append(i)
            else:
                cevaplar[f"soru{i}"] = harf

        if eksik:
            hatalar.append(f"{row_idx}. satırda eksik/yanlış cevaplar: {', '.join(map(str, eksik))}")
            continue

        # Mevcut puanlama ve alan hesaplarını kullan
        puanlar, genel_A, genel_G = puan_hesapla(cevaplar)
        alanlar = hesapla_johari_alanlari(puanlar)

        sonuclar.append({
            "yapan": str(yapan or "").strip(),
            "test_edilen": str(edilen or "").strip(),
            "puanlar": puanlar,
            "genel_A": genel_A,
            "genel_G": genel_G,
            "alanlar": alanlar
        })
    return sonuclar, hatalar
# --- Excel şablonu indirme (DOĞRU: /excel-sablon) ---
@app.route("/excel-sablon", endpoint="excel_sablon")
def excel_sablon():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cevaplar"
    headers = ["Yapan", "Test Edilen"] + [f"S{i}" for i in range(1, 49)]
    ws.append(headers)
    # örnek satır
    ws.append(["Ali", "Veli"] + ["A"] * 48)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="johari_sablon.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# --- Excel yükleme ve hesaplama (DOĞRU: /excel-yukle) ---
@app.route("/excel-yukle", methods=["GET", "POST"], endpoint="excel_yukle")
def excel_yukle():
    if request.method == "GET":
        return render_template("excel_yukle.html")

    f = request.files.get("dosya")
    if not f or not f.filename.lower().endswith(".xlsx"):
        flash("Lütfen .xlsx uzantılı Excel dosyası yükleyin.", "error")
        return redirect(url_for("excel_yukle"))

    try:
        sonuclar, hatalar = _exceli_oku_ve_hesapla(f)
    except Exception as e:
        flash(f"Dosya okunamadı: {e}", "error")
        return redirect(url_for("excel_yukle"))

    if hatalar:
        for h in hatalar:
            flash(h, "error")
        if not sonuclar:
            return redirect(url_for("excel_yukle"))

    session["excel_sonuc_list"] = sonuclar  # sadece bu oturumda
    return redirect(url_for("excel_sonuc"))

@app.route("/excel-sonuc", endpoint="excel_sonuc")
def excel_sonuc():
    lst = session.get("excel_sonuc_list")
    if not lst:
        flash("Önce bir Excel yükleyin.", "error")
        return redirect(url_for("excel_yukle"))
    # Bu sayfayı gösterecek basit bir tablo şablonun yoksa,
    # geçici olarak JSON gibi gösterebiliriz:
    # return {"sonuclar": lst}
    return render_template("excel_sonuc.html", liste=lst)

@app.route("/excel-sonuc-indir", endpoint="excel_sonuc_indir")
def excel_sonuc_indir():
    lst = session.get("excel_sonuc_list")
    if not lst:
        abort(404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sonuçlar"
    ws.append([
        "Yapan", "Test Edilen",
        "Genel A", "Genel G",
        "Açık %", "Kör %", "Gizli %", "Bilinmeyen %"
    ])

    for r in lst:
        alan = r["alanlar"]
        ws.append([
            r["yapan"], r["test_edilen"],
            r["genel_A"], r["genel_G"],
            alan["acik_yuzde"], alan["kor_yuzde"],
            alan["gizli_yuzde"], alan["bilinmeyen_yuzde"]
        ])

    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="johari_excel_sonuc.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ================== ROUTELAR ==================
@app.route("/")
def giris():
    # Modern hero tasarımlı giriş
    return render_template("giris.html")

@app.route("/index", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        yapan = request.form.get("yapan", "").strip()
        test_edilen = request.form.get("test_edilen", "").strip()
        if request.form.get("kendim_icin") == "on":
            test_edilen = yapan

        # Sorular yükle
        from sorular import sorular

        # Oturuma temel bilgiler
        session.permanent = True
        session['yapan'] = yapan
        session['test_edilen'] = test_edilen
        session['cevap_id'] = str(uuid.uuid4())

        # Test sayfası
        return render_template("test.html", yapan=yapan, test_edilen=test_edilen, sorular=sorular)
    return render_template("index.html")  # mevcut sayfan varsa çalışmaya devam etsin

@app.route("/indir/sonuclar")
def indir_sonuc():
    admin_key = os.environ.get("ADMIN_PASS")
    key = request.args.get("key")
    if not admin_key or key != admin_key:
        abort(403)
    path = os.path.join(current_app.root_path, "data", "sonuclar.xlsx")
    if not os.path.exists(path):
        abort(404, description="Henüz kayıt yok.")
    return send_file(path, as_attachment=True, download_name="johari_tum_sonuclar.xlsx")


# === GİZLİ RAPOR/EXCEL İNDİRME (Admin korumalı) ===
from flask import send_file, abort, current_app
from werkzeug.utils import safe_join
import shutil, time

@app.route("/indir/benim-sonucum")
def indir_benim_sonucum():
    data = session.get("sonuc")
    if not data:
        return redirect(url_for("index"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Benim Sonucum"
    ws.append([
        "Tarih", "Yapan", "Test Edilen",
        "A1", "A2", "G1", "G2", "Genel A", "Genel G",
        "Açık", "Kör", "Gizli", "Bilinmeyen",
        "Açık (%)", "Kör (%)", "Gizli (%)", "Bilinmeyen (%)"
    ])

    alanlar = data["alanlar"]
    puanlar = data["puanlar"]
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        data["yapan"], data["test_edilen"],
        puanlar["A1"], puanlar["A2"], puanlar["G1"], puanlar["G2"],
        data["genel_A"], data["genel_G"],
        round(alanlar["acik"], 2), round(alanlar["kor"], 2),
        round(alanlar["gizli"], 2), round(alanlar["bilinmeyen"], 2),
        f"%{alanlar['acik_yuzde']}", f"%{alanlar['kor_yuzde']}",
        f"%{alanlar['gizli_yuzde']}", f"%{alanlar['bilinmeyen_yuzde']}"
    ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    fname = f"johari_{data['yapan']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route("/eposta-gonder", methods=["POST"])
def eposta_gonder():
    yapan = request.form.get("yapan")
    eposta = request.form.get("eposta")
    grafik_path = request.form.get("grafik_path")
    yorum = request.form.get("yorum")
    alanlar = {
        "acik_yuzde": float(request.form.get("alanlar_acik", 0)),
        "kor_yuzde": float(request.form.get("alanlar_kor", 0)),
        "gizli_yuzde": float(request.form.get("alanlar_gizli", 0)),
        "bilinmeyen_yuzde": float(request.form.get("alanlar_bilinmeyen", 0))
    }
    mail_gonder(eposta, yapan, grafik_path, alanlar, yorum)
    return redirect(url_for("sonuc_get"))


@app.route("/sonuc", methods=["POST"])
def sonuc_post():
    yapan = request.form["yapan"]
    test_edilen = request.form["test_edilen"]
    cevaplar = {k: v for k, v in request.form.items() if k.startswith("soru")}
    puanlar, genel_A, genel_G = puan_hesapla(cevaplar)
    alanlar = hesapla_johari_alanlari(puanlar)

    # ÜSTTE Açık & Kör olacak şekilde grafik
    grafik_path = ciz_grafik_duzenli(alanlar, yapan)

    # Yapay zeka yorumu
    yorum = yapay_zeka_yorumla(yapan, test_edilen, alanlar)

    # Excel'e KAYIT (artık data/sonuclar.xlsx)
    kaydet_excel(yapan, test_edilen, puanlar, genel_A, genel_G, alanlar)

    # PRG (Post/Redirect/Get) için session'a özet koy
    session["sonuc"] = {
        "yapan": yapan,
        "test_edilen": test_edilen,
        "puanlar": puanlar,
        "genel_A": genel_A,
        "genel_G": genel_G,
        "grafik_path": grafik_path,
        "alanlar": alanlar,
        "yorum": yorum
    }
    return redirect(url_for("sonuc_get"))

@app.route("/sonuc", methods=["GET"])
def sonuc_get():
    data = session.get("sonuc")
    if not data:
        return redirect(url_for("index"))
    return render_template("sonuc.html", **data)
@app.route("/__selftest")
def __selftest():
  
    harfler = ["C","B","A","C","D","B","A","D","D","D","A","B","D","C","A","B","A","C","A","D","D","C","C","C","D","B","A","D","C","C","D","D","D","C","A","B","D","D","B","C","B","D","B","D","D","C","D","C"]
    cevaplar = {f"soru{i+1}": harfler[i] for i in range(48)}

    puanlar, A, G = puan_hesapla(cevaplar)
    alan = hesapla_johari_alanlari(puanlar)
    return {
        "gruplanan_puanlar": puanlar,  # A1, A2, G1, G2
        "A": A, "G": G,
        "yuzdeler": {
            "acik": alan["acik_yuzde"],
            "kor": alan["kor_yuzde"],
            "gizli": alan["gizli_yuzde"],
            "bilinmeyen": alan["bilinmeyen_yuzde"],
        }
    }


def puan_hesapla(cevaplar):
    # Madde grupları — 39 G2'de
    G1 = [1, 4, 6, 14, 16, 24, 26, 34, 36, 40, 46, 47]
    G2 = [3, 9, 12, 18, 21, 28, 30, 31, 37, 39, 41, 44]
    A1 = [2, 5, 7, 13, 17, 19, 23, 25, 27, 29, 32, 35]
    A2 = [8, 10, 11, 15, 20, 22, 33, 38, 42, 43, 45, 48]

    # Excel ölçeği
    harf_puanlari = {"A": 4, "B": 3, "C": 2, "D": 1}

    puanlar = {"G1": 0, "G2": 0, "A1": 0, "A2": 0}

    for key, secenek in cevaplar.items():
        sn = int(key.replace("soru", ""))
        val = harf_puanlari.get(str(secenek).upper(), 0)

        if sn in G1:
            puanlar["G1"] += val
        elif sn in G2:            # G2 ters
            puanlar["G2"] += (5 - val)
        elif sn in A1:
            puanlar["A1"] += val
        elif sn in A2:            # A2 ters
            puanlar["A2"] += (5 - val)

    # A ve G Excel'deki gibi 0..48 aralığında
    genel_A = round((puanlar["A1"] + puanlar["A2"]) / 2, 2)
    genel_G = round((puanlar["G1"] + puanlar["G2"]) / 2, 2)
    return puanlar, genel_A, genel_G


def hesapla_johari_alanlari(puanlar):
    A = (puanlar["A1"] + puanlar["A2"]) / 2  # 0..48
    G = (puanlar["G1"] + puanlar["G2"]) / 2  # 0..48

    # 48^2 = 2304 -> Excel formülü birebir
    alan_acik = (G * A) / 2304
    alan_kor = ((48 - G) * A) / 2304
    alan_gizli = (G * (48 - A)) / 2304
    alan_bilinmeyen = ((48 - G) * (48 - A)) / 2304

    return {
        "acik": alan_acik, "kor": alan_kor, "gizli": alan_gizli, "bilinmeyen": alan_bilinmeyen,
        "acik_yuzde": round(alan_acik * 100, 2),
        "kor_yuzde": round(alan_kor * 100, 2),
        "gizli_yuzde": round(alan_gizli * 100, 2),
        "bilinmeyen_yuzde": round(alan_bilinmeyen * 100, 2),
        "G": round(G, 2), "A": round(A, 2)
    }





def ciz_grafik_duzenli(alanlar, yapan_adi):
    G = max(0, min(48, alanlar["G"]))
    A = max(0, min(48, alanlar["A"]))

    fig, ax = plt.subplots(figsize=(8, 8))
    ax.set_xlim(0, 48)
    ax.set_ylim(0, 48)
    ax.set_aspect("equal")
    ax.axis("off")

 
    rects = [
        ((0, 48 - A),       G,      A,      "Açık",        alanlar['acik_yuzde']),
        ((G, 48 - A),       48 - G, A,      "Kör",         alanlar['kor_yuzde']),
        ((0, 0),            G,      48 - A, "Gizli",       alanlar['gizli_yuzde']),
        ((G, 0),            48 - G, 48 - A, "Bilinmeyen",  alanlar['bilinmeyen_yuzde']),
    ]
    colors = ["#b8e994", "#f8c291", "#82ccdd", "#d1d8e0"]

    for i, (xy, w, h, ad, yuzde) in enumerate(rects):
        ax.add_patch(patches.Rectangle(xy, w, h, alpha=0.7, ec="#2f3640", fc=colors[i], lw=1.5))
        if w > 2 and h > 2:
            ax.text(xy[0] + w/2, xy[1] + h/2, f"{ad}\n%{yuzde}", ha="center", va="center", fontsize=12, weight="bold")

    # Bölücü çizgiler
    ax.axhline(y=48 - A, color="black", linewidth=1.5)
    ax.axvline(x=G, color="black", linewidth=1.5)

    # Dosya
    kisi = slugify(yapan_adi)
    dosya_adi = f"johari_{kisi}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    tam_yol = os.path.join(GRAFIK_DIR, dosya_adi)
    fig.tight_layout()
    fig.savefig(tam_yol, dpi=150)
    plt.close(fig)
    return f"grafik/{dosya_adi}"

import os, requests

import os, requests, hashlib
from flask import session


import os, requests, hashlib
from flask import session
=
import os, requests, hashlib, random, time
from flask import session

def _profil_etiketi(a, k, g, b):
   
    vals = {"acik": a, "kor": k, "gizli": g, "bilinmeyen": b}
    srt = sorted(vals.items(), key=lambda x: x[1], reverse=True)
    birinci, ikinci = srt[0][0], srt[1][0]
    return f"{birinci}+{ikinci}", birinci, ikinci

def _ton_sec(seed_int):
    tonlar = [
        "sakin", "analitik", "sıcak", "ilham verici",
        "geribildirim odaklı", "çözüm odaklı", "yansıtıcı", "net ve yalın"
    ]
    random.seed(seed_int)
    return random.choice(tonlar)

def _ipucu_bankasi():
   
    return {
        "acik": [
            "Toplantı öncesi kısa bir bağlam notu paylaşmak, görünürlüğü doğal biçimde artırır.",
            "Karar sonrasında iki cümlelik mini-özet, ekipte ortak anlayışı hızla pekiştirir.",
            "Küçük başarıları görünür kılmak, güven duygusunu süreklileştirir."
        ],
        "kor": [
            "İki yönlü geri bildirim ritmi kurmak, algı-niyet farkını hızla kapatır.",
            "Varsayımları yüksek sesle dile getirmek, sürprizleri azaltır.",
            "Ekipten örnek davranış-geribildirim örnekleri istemek, kör noktaları aydınlatır."
        ],
        "gizli": [
            "Kişisel çalışma tercihlerini iki-üç net cümleyle yazmak, beklenti yönetimini kolaylaştırır.",
            "Öncelikleri küçük notlarla paylaşmak, gereksiz gerginlikleri azaltır.",
            "Sınırları belirginleştirmek, ilişki kalitesini hızla yükseltir."
        ],
        "bilinmeyen": [
            "İki haftalık küçük pilot denemeler, yeni becerileri risksiz ortamda test etmenizi sağlar.",
            "Bir ‘öğrenme günlüğü’ ile keşifleri görünür kılmak, ilerlemeyi hızlandırır.",
            "Farklı rollerde kısa rotasyonlar, gizli potansiyelleri ortaya çıkarır."
        ]
    }

def _varyasyon_ifadeleri():
    # Dilde küçük varyasyonlar için eşanlam bankası
    return {
        "gosteriyor": ["gösteriyor", "işaret ediyor", "ortaya koyuyor", "resmediyor"],
        "odak": ["odağı", "önceliği", "merkezi", "ağırlık noktası"],
        "olabilir": ["olabilir", "mümkün", "uygun görünüyor", "yerinde olacaktır"],
        "destekler": ["destekler", "pekiştirir", "güçlendirir", "kolaylaştırır"]
    }

def _sec(bank, rng):
    return rng.choice(bank)

def _yonlendirme_cumlesi(birinci, rng, bank):
    # Baskın alana göre bir öneri cümlesi seç
    harita = {
        "acik": bank["acik"],
        "kor": bank["kor"],
        "gizli": bank["gizli"],
        "bilinmeyen": bank["bilinmeyen"]
    }
    return _sec(harita[birinci], rng)

def yapay_zeka_yorumla(yapan, test_edilen, alanlar):
    # Tohum: kişi + skorlar + oturum cevabı + zaman tuzu (tekrarı kırmak için hafif varyasyon)
    seed_src = (
        f"{yapan}|{test_edilen}|"
        f"{alanlar['acik_yuzde']}-{alanlar['kor_yuzde']}-"
        f"{alanlar['gizli_yuzde']}-{alanlar['bilinmeyen_yuzde']}|"
        f"{session.get('cevap_id','')}|{int(time.time())//120}"  # her ~2 dakikada bir farklılaşsın
    )
    seed_int = int(hashlib.sha256(seed_src.encode('utf-8')).hexdigest(), 16)
    rng = random.Random(seed_int)

    ton = _ton_sec(seed_int)
    a = float(alanlar['acik_yuzde'])
    k = float(alanlar['kor_yuzde'])
    g = float(alanlar['gizli_yuzde'])
    b = float(alanlar['bilinmeyen_yuzde'])

    profil, birinci, ikinci = _profil_etiketi(a, k, g, b)
    vary = _varyasyon_ifadeleri()
    ipuclar = _ipucu_bankasi()

    # Profil bazlı vurgu cümleleri
    profil_cumleleri = {
        "acik+kor": "Geniş bir görünürlükle birlikte zaman zaman algı-niyet farkları ortaya çıkabiliyor.",
        "acik+gizli": "Paylaşım yüksek; ancak kişisel tercihlerin bir kısmı içeride tutuluyor.",
        "acik+bilinmeyen": "Görünürlük iyi; keşif alanında hâlâ büyüme potansiyeli var.",
        "kor+acik": "Algı farklılıkları belirgin; yine de görünürlük zemini güçlü.",
        "kor+gizli": "Hem beklentiler hem de algılar kapalı kalmaya meyilli; güvenli alanlar kurmak önemli.",
        "kor+bilinmeyen": "Geribildirimin sınırlı kaldığı ve keşfin düşük olduğu bir görüntü var.",
        "gizli+acik": "İletişim iyi; ancak kişisel sınırlar ve tercihlerin görünürlüğü artabilir.",
        "gizli+kor": "Paylaşım kısıtlı; algı farkı riski yükseliyor.",
        "gizli+bilinmeyen": "İçe dönük kalma eğilimi ile belirsiz alan birlikte ilerliyor.",
        "bilinmeyen+acik": "Görünürlük sağlam; keşfe ayrılacak küçük adımlar hızlı fayda getirebilir.",
        "bilinmeyen+kor": "Geribildirim ve deney alanı ikisi de sınırlı; ritim yaratmak kritik.",
        "bilinmeyen+gizli": "Keşif ve paylaşım düşük; küçük, güvenli denemelerle başlamak en doğrusu."
    }
    ana_mesaj = profil_cumleleri.get(f"{birinci}+{ikinci}", "Dört alanın dengesi kişiye özel bir harita ortaya koyuyor.")

    # Online API için daha kişiselleştirilmiş, profil tabanlı talimat
    prompt = f"""
Johari sonucu (Türkçe, serbest metin, {ton} ton):
Kişi: {test_edilen}
Oranlar: Açık %{a:.2f}, Kör %{k:.2f}, Gizli %{g:.2f}, Bilinmeyen %{b:.2f}
Profil: {profil} (baskın: {birinci}, ikincil: {ikinci})

Görev:
- 2 veya 3 paragraf yaz.
- Başlık, madde işareti, emoji yok.
- İlk paragrafta tabloyu {rng.choice(vary['gosteriyor'])}; ikinci paragrafta eylem önerilerini doğal akışta yedir.
- Öneriler profil ile tutarlı olsun: '{birinci}' ve '{ikinci}' alanlarına odaklan.
- Küçük ama uygulanabilir adımlar öner (kural: bağlam notu/varsayım görünür kılma/mini pilot/öğrenme günlüğü gibi).
- Dili sade, akıcı; klişe kaçın.
- 180–260 kelime aralığı.

Bağlam cümlesi: {ana_mesaj}
"""

    api_key = os.environ.get("OPENROUTER_API_KEY", "").strip()
    if api_key:
        base_url = os.environ.get("OPENROUTER_URL", "https://openrouter.ai/api/v1")
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": os.environ.get("OR_REFERER", "https://example.com"),
            "X-Title": "Johari Test Platformu"
        }
        # Birkaç model + farklı sıcaklıklarla deneyerek çeşitlilik arttır
        candidates = [
            ("openrouter/auto", 0.95, 0.9),
            ("mistralai/mistral-small:free", 1.05, 0.92),
            ("mistralai/mistral-nemo:free", 1.0, 0.95),
            ("qwen/qwen2.5-7b-instruct:free", 1.1, 0.9),
            ("meta-llama/llama-3.1-8b-instruct:free", 1.05, 0.9),
        ]

        rng.shuffle(candidates)  # her çağrıda farklı sıralama
        for model, temp, topp in candidates:
            try:
                r = requests.post(
                    f"{base_url}/chat/completions",
                    headers=headers,
                    json={
                        "model": model,
                        "messages": [
                            {"role": "system", "content": "Profesyonel, doğal, kişiye özgü Türkçe anlatı yaz. Tekrar ve şablondan kaçın."},
                            {"role": "user", "content": prompt}
                        ],
                        "max_tokens": 900,
                        "temperature": float(temp),
                        "top_p": float(topp),
                        "presence_penalty": 0.55,
                        "frequency_penalty": 0.45
                    },
                    timeout=25
                )
                if r.status_code == 200:
                    txt = (r.json().get("choices") or [{}])[0].get("message", {}).get("content", "") or ""
                    # Tekrarlayan çıktıyı kır: öncekiyle aynıysa küçük bir tuz ekleyip yeniden dener
                    last_hash = session.get("last_comment_hash")
                    cur_hash = hashlib.md5(txt.encode("utf-8")).hexdigest()
                    if last_hash and last_hash == cur_hash:
                        prompt += f"\nNot: Aynı ifadelerden kaçın; farklı metafor ve cümle yapıları kullan."
                        continue
                    session["last_comment_hash"] = cur_hash
                    if len(txt.split()) >= 160:
                        return txt.strip()
            except Exception:
                continue

    # Offline (API yoksa) – profil ve seed tabanlı yerel üretim
    return _yerel_serbest_yorum(test_edilen, a, k, g, b, ton, profil, rng, vary, ipuclar)

def _yerel_serbest_yorum(ad, a, k, g, b, ton, profil=None, rng=None, vary=None, ipuclar=None):
    # Seed’li çeşitlilik: eşanlam + profil odaklı cümleler + ipuçları
    if rng is None:
        rng = random.Random(int(hashlib.md5(f"{ad}-{a}-{k}-{g}-{b}".encode()).hexdigest(), 16))
    vary = vary or _varyasyon_ifadeleri()
    ipuclar = ipuclar or _ipucu_bankasi()

    # Basit seviye metrikleri
    def seviye(x):
        return "çok yüksek" if x >= 40 else "yüksek" if x >= 28 else "orta" if x >= 18 else "düşük"

    # Baskın alan
    vals = {"acik": a, "kor": k, "gizli": g, "bilinmeyen": b}
    birinci = max(vals, key=vals.get)
    ikinci = max({k:v for k,v in vals.items() if k != birinci}, key=lambda z: vals[z])

    giris = (
        f"{ad} için ortaya çıkan tablo, güçlü yanlarla gelişime açık noktaların birlikte "
        f"{rng.choice(vary['gosteriyor'])}. Açık alanın {seviye(a)} düzeyi, niyet ve çalışma tarzının "
        f"çoğunlukla anlaşılır olduğunu düşündürür; bu durum ekip içi güveni {rng.choice(vary['destekler'])}. "
        f"Buna karşılık kör alanın {seviye(k)} görünmesi, algı ile niyetin zaman zaman farklılaşabildiğine işaret eder. "
        f"Gizli alanın {seviye(g)} olması, uygun bağlam sağlandığında ilişkilerin daha rahat akabileceğini; "
        f"bilinmeyen alanın {seviye(b)} düzeyi ise küçük denemelerin yeni becerileri hızla görünür kılabileceğini {rng.choice(vary['gosteriyor'])}."
    )

    ipucu_cumleleri = []
    ipucu_cumleleri.append(_yonlendirme_cumlesi(birinci, rng, ipuclar))
    if rng.random() < 0.7:
        ipucu_cumleleri.append(_yonlendirme_cumlesi(ikinci, rng, ipuclar))
    if rng.random() < 0.5:
        # Üçüncü bir mikro-öneri; en düşük alandan seç
        ucuncu = min(vals, key=vals.get)
        ipucu_cumleleri.append(_yonlendirme_cumlesi(ucuncu, rng, ipuclar))

    gelisim = (
        f"Bu resimde {rng.choice(vary['odak'])}, açık alanı bilinçli paylaşımla büyütmek ve kör alanı "
        f"istikrarlı geri bildirimle daraltmak {rng.choice(vary['olabilir'])}. "
        f"{' '.join(ipucu_cumleleri)} "
        f"Kısa döngülerle yapılan değerlendirmeler, hız yerine ritim kurmayı sağlar; böylece güçlü yanlar belirginleşirken "
        f"gelişime açık taraflar zorlanmadan dönüşür."
    )

    if rng.random() < 0.5:
        kapanis = (
            f"Zaman içinde bu dört alanın dengesi değiştikçe iletişim sadeleşir ve karar süreçleri netleşir. "
            f"{ad} için öneri: düzenli görünürlük, merakla dinleme ve küçük pilotlar. "
            f"Bu üçlü, doğal bir öğrenme hattı oluşturur."
        )
        return f"{giris}\n\n{gelisim}\n\n{kapanis}"
    else:
        return f"{giris}\n\n{gelisim}"



def kaydet_excel(yapan, test_edilen, puanlar, genel_A, genel_G, alanlar):
    os.makedirs("data", exist_ok=True)
    dosya_yolu = os.path.join("data", "sonuclar.xlsx")
    if not os.path.exists(dosya_yolu):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sonuçlar"
        ws.append([
            "Tarih", "Yapan", "Test Edilen",
            "A1", "A2", "G1", "G2", "Genel A", "Genel G",
            "Açık", "Kör", "Gizli", "Bilinmeyen",
            "Açık (%)", "Kör (%)", "Gizli (%)", "Bilinmeyen (%)"
        ])
    else:
        wb = load_workbook(dosya_yolu)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        yapan, test_edilen,
        puanlar["A1"], puanlar["A2"], puanlar["G1"], puanlar["G2"],
        genel_A, genel_G,
        round(alanlar["acik"], 2), round(alanlar["kor"], 2),
        round(alanlar["gizli"], 2), round(alanlar["bilinmeyen"], 2),
        f"%{alanlar['acik_yuzde']}", f"%{alanlar['kor_yuzde']}",
        f"%{alanlar['gizli_yuzde']}", f"%{alanlar['bilinmeyen_yuzde']}"
    ])
    wb.save(dosya_yolu)




def mail_gonder(eposta, yapan, grafik_path, alanlar, yorum):
    try:
        msg = Message(
            subject="Johari Test Sonucunuz ve Tavsiyeler",
            recipients=[eposta]
        )
        msg.body = (
            f"Merhaba {yapan},\n\n"
            "Johari Test Sonuçlarınız:\n"
            f"Açık Alan: %{alanlar['acik_yuzde']}\n"
            f"Kör Alan: %{alanlar['kor_yuzde']}\n"
            f"Gizli Alan: %{alanlar['gizli_yuzde']}\n"
            f"Bilinmeyen Alan: %{alanlar['bilinmeyen_yuzde']}\n\n"
            f"Yapay Zeka Yorumu:\n{yorum}\n\n"
            "Sevgiler,\nJohari Testi"
        )
        with app.open_resource(os.path.join("static", grafik_path)) as fp:
            msg.attach("johari_sonuc.png", "image/png", fp.read())
        mail.send(msg)
    except Exception as e:
        print(f"[HATA] Mail gönderilemedi: {str(e)}")

@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

if __name__ == "__main__":
    app.run(debug=True)
